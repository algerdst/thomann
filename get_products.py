import glob
import json
import os
import time
import openpyxl

import requests
from bs4 import BeautifulSoup

headers = {
  'authority': 'www.thomann.de',
  'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
  'accept-language': 'ru,en;q=0.9',
  'cache-control': 'max-age=0',
  'cookie': 'thomann_settings=6f16d8a5-2c56-4b3b-bb9a-e89c22a138ad; _fbp=fb.1.1710315334492.1744237242; FPID=FPID2.2.O8auqg5OeqiUtLpHqIZxjkWRJ7VgSUSFVxqANzMmLzE%3D.1710315334; _gid=GA1.2.452134362.1710315337; _gcl_au=1.1.568821786.1710315337; _scid=b64fa071-5e8c-4c4f-ae64-0157ac07eb83; _tt_enable_cookie=1; _ttp=0i1_JK6HRfpM2CBi7TApwl85lCo; _pin_unauth=dWlkPVlqSXhPRFV4WTJRdFpXTmxPUzAwWVRGbUxUbGlNVFF0TkRjeVpEUXpORFEzTkRKbA; _sctr=1%7C1710277200000; uslk_umm_27718_s=ewAiAHYAZQByAHMAaQBvAG4AIgA6ACIAMQAiACwAIgBkAGEAdABhACIAOgB7AH0AfQA=; sid=3d5e86a18d887a8fe4580c966f47110b; FPLC=KJ24cHA1tJYOZJvPqAS7SI2uL2eMQwey06rR2O5z3MPVDiYgzbfcBTUmB6sbkGmxh2Eh8XYodALYOm1pvXxUbdsoQqp74oy1PWdtvJYOD4EiVQEYFP5GwvIoQH91rg%3D%3D; _gat_UA-6023113-1=1; _ga_CQB6MP7VD2=GS1.1.1710414195.10.1.1710414525.59.0.0; _scid_r=b64fa071-5e8c-4c4f-ae64-0157ac07eb83; _ga_QNTG1E3BFT=GS1.1.1710414193.10.1.1710414525.0.0.0; _ga=GA1.2.1454382977.1710315334; _br_uid_2=uid%3D9720173930277%3Av%3D15.0%3Ats%3D1710315320588%3Ahc%3D93',
  'referer': 'https://www.thomann.de/intl/harley_benton_electric_guitars.html?ls=25&pg=1',
  'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "YaBrowser";v="24.1", "Yowser";v="2.5"',
  'sec-ch-ua-mobile': '?0',
  'sec-ch-ua-platform': '"Windows"',
  'sec-fetch-dest': 'document',
  'sec-fetch-mode': 'navigate',
  'sec-fetch-site': 'same-origin',
  'sec-fetch-user': '?1',
  'upgrade-insecure-requests': '1',
  'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36'
}


def get_product():
    file = []
    path = os.getcwd()
    for filename in glob.glob(os.path.join(path, '*.xlsx')):
        file.append(filename)
    filename = file[0]
    book = openpyxl.load_workbook(filename)
    sheet=book.active
    with open('links.json', 'r', encoding='utf-8') as file:
        links_dict = json.load(file)
    row=2
    count=0
    with open('отработанные ссылки.txt', 'r', encoding='utf-8') as file:
        for link in file:
            link=link.replace('\n', '')
            if link in links_dict:
                del links_dict[link]
    for link in links_dict:
        category=links_dict[link]
        response = requests.get(link, headers=headers)
        time.sleep(1)
        soup = BeautifulSoup(response.text, 'lxml')
        try:
            id=soup.findAll('div', 'keyfeature__wrapper')[1].findAll('span')[1].text
            name = soup.find('h1').text
            price = soup.find('div', class_='price').text.replace('\n', '').strip()
            try:
                price=price.split('$')[1]
                price=price.replace('€', '').strip()
            except:
                price = price.replace('€', '').strip()
            description = f"<p>Гитара {name}  под заказ из Европы.</p> <p>Магазины Musicstore и Thomann.</p> <p>Доставка от 4-8 недель с момента отправки, по срокам доставки нужно уточнять дополнительно, может меняться в зависимости от наличия.</p> <p>🚛 Oтпpaвлю по всeй Poccии удoбным для вас способом</p> <p>✅ Можете выбpaть любую дpугую гитаpу, электрогитару или акустическую гитару под заказ, даже если ее нет в моих объявлениях, просто напишите полное название.</p> <p>✔ ГAРAHTИЯ НА ТOВАР</p> <p>📨 Пишите</p> <p>📞 Звоните</p> <p>👨‍🔧Без выходных</p>"
            img = soup.findAll('img', class_='navigator__item-image')
            images_list=[]
            for i in img:
                if len(images_list)<10:
                    images_list.append(i)
            if len(images_list)>1:
                images = " | ".join([i['src'].replace('80x80', '600x600') for i in images_list])
            else:
                try:
                    images= soup.find('div', class_='product-media-gallery').find('img', class_='spotlight__item-image')['src']
                except:
                    images='-'
        except:
            continue
        sheet.cell(column=2, row=row).value=id
        sheet.cell(column=4, row=row).value=price
        sheet.cell(column=12, row=row).value=name
        sheet.cell(column=13, row=row).value=category
        sheet.cell(column=14, row=row).value=description
        sheet.cell(column=19, row=row).value=images
        row+=1
        book.save(filename)
        count+=1
        with open('отработанные ссылки.txt', 'a', encoding='utf-8') as file:
            file.write(link + '\n')
        print(f'Записал товар {name} в таблицу, осталось собрать товаров - {len(links_dict)-count} ')
    book.close()


get_product()